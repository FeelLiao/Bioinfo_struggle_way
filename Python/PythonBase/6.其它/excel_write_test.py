#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May  7 12:53:51 2020

@author: feel-liao
"""
# update the new pop data of some countries
import openpyxl
from openpyxl.utils import column_index_from_string
wb=openpyxl.load_workbook("/home/feel-liao/Downloads/population.xlsx")
pop_data=wb.get_sheet_by_name("Data")
# the updated data of countries
update_pop={
    "Chile":8932435321,
    "Benin":293374832,
    "Belarus":244643236,
    "Guinea-Bissau":353354632
    }
# loop through the rows and update the pop data
for row_num in range(5,pop_data.max_row+1):
    country_name=pop_data.cell(row=row_num,column=1).value
    if country_name in update_pop:
        print(pop_data.cell(row=row_num,
                        column=column_index_from_string("E")).value)
        pop_data.cell(row=row_num,
column=column_index_from_string("E")).value=update_pop[country_name]
        print(pop_data.cell(row=row_num,
                        column=column_index_from_string("E")).value)
        print("------------")
wb.save("/home/feel-liao/Python/excel/updatewb.xlsx")




        
        
        