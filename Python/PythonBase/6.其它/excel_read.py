#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed May  6 08:22:58 2020

@author: feel-liao
"""
import openpyxl
wb=openpyxl.load_workbook("/home/feel-liao/Downloads/粗灰分测定.xlsx")
# workbook worksheet row,column,cell
# Getting sheets from the workbook
# print(wb.sheetnames)
mySheet=wb.create_sheet("mysheet")
# creat a new sheet named mysheet
sheet3=wb.get_sheet_by_name("3")
# variables sheet3 set to sheet 3 in the workbook
cell_sheet3_A1=sheet3["A1"]
# print(cell_sheet3_A1.value) 
# read the value of cell A1
"""
print("row {}, column {} is {}".format(cell_sheet3_A1.row,
cell_sheet3_A1.column,cell_sheet3_A1.value))
print("cell {} is {}".format(cell_sheet3_A1.coordinate,
                                       cell_sheet3_A1.value))
print(sheet3.cell(row=1,column=1))
print(sheet3.cell(row=1,column=1).value)
print the information of cell A1(cell_sheet3_A1)
"""
"""
cell_range=sheet3["B3:C5"]
for row_of_cell_obj in cell_range:
    for cell_obj in row_of_cell_obj:
        print(cell_obj.coordinate,cell_obj.value)
    print("-----end of row------")
get part of the sheet data
"""
# print("{}*{}".format(sheet3.max_row,sheet3.max_column))
# get to know the max row and column of this sheet
"""
from openpyxl.utils import get_column_letter,column_index_from_string
print(get_column_letter(30))
print(column_index_from_string("AB"))
change the letters into number 
 """      




