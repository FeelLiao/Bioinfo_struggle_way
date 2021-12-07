#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri May  8 18:10:57 2020

@author: feel-liao
"""
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
wb=openpyxl.Workbook()
# Font
Font_set=wb.create_sheet(title="Font")
# default 11pt,Calibri
italic24Font=Font(size=24,italic=True)
Font_set["A3"].font=italic24Font
Font_set["A3"]="Font test "
Font_set['B3']="italic24Font"
boldRedFont=Font(name="Times New Roman",bold=True,color=colors.RED)
Font_set["A2"].font=boldRedFont
Font_set["A2"]="Font test"
Font_set["B2"]="boldRedFont"
# Formulas
Formulas=wb.create_sheet(title="Formulas")
Formulas["A1"]=200
Formulas["A2"]=300
Formulas["B1"]="Total"
Formulas["B2"]="=sum(A1:A2)"
# setting row height and column width
dimensions=wb.create_sheet(title="dementions")
dimensions["A1"]="Tall row"
dimensions.row_dimensions[1].height=70
dimensions["B2"]="Wide column"
dimensions.column_dimensions["B"].width=20
# Merging cells
merged=wb.create_sheet("merged")
merged.merge_cells("A1:D3")
merged["A1"]="merged cells"
# unmerging cells
# unmerged.merge_cells("A1:D3")
wb.save("/home/feel-liao/Python/excel/excel_format.xlsx")



