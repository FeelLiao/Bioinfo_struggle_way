#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May  7 09:29:19 2020

@author: feel-liao
"""
import openpyxl
from openpyxl.utils import get_column_letter
wb=openpyxl.Workbook()
sheet=wb.active
# chage the title of the sheet
sheet.title="range names"
wb.create_sheet(index=1,title="List")
# save the new workbook
# wb.save("/home/feel-liao/Python/excel/test1.xlsx")
# delete the sheet 
# wb.remove_sheet(wb.get_sheet_by_name("diagram"))
# print(wb.get_sheet_names())
# write value to cells
range_names_sheet=wb.get_sheet_by_name("range names")
list_sheet=wb.get_sheet_by_name("List")
data_sheet=wb.create_sheet(title="data")
# range_names_sheet["A1"]="hello python"
# print(range_names_sheet["A1"].value)
# for row in range(1,40):
    # range_names_sheet.append(range(17))
    # add 0 to 16 to each row until 39
"""
rows=[
      ["number","batch 1","batch 2"],
      [1,2,3],
      [4,5,6],
      [7,8,9],
      ]
for row in rows:
    list_sheet.append(row)
    
add the data according to the format of the list
"""
"""
for row in range(5,30):
    for col in range(15,30):
        data_sheet.cell(column=col,row=row,value=get_column_letter(col))
"""
wb.save("/home/feel-liao/Python/excel/test1.xlsx")


